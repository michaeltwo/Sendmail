# -*- coding: utf-8 -*-

#程序名称：自动发送邮件脚本
#程序描述：实现从本地路径读取excel文件内容，并绘制HTML表格，将内容写入邮件正文区发送
#编写时间：2017.11.15
#作者：MIC
#解释环境：python3.X


#引入excel文档相关库
import xlrd
import os,time
from openpyxl import Workbook
from openpyxl import load_workbook
#---------------------------------------------------
wb = load_workbook(filename = 'C:\\File\\file.xlsx')
sheet_ranges = wb['Sheet2']
ws = wb.get_sheet_by_name('Sheet2')
total = ws.max_row
#----------------------------------------------------
#引入邮件相关库
import smtplib
import win32com.client as win32
from email.mime.text import MIMEText
from email.header import Header
import base64
#处理从excel中读取的float类型数据的类
#目前集成两种处理：（1）float到int型的转换（2）float到str型的转换，后续有需要可以增加方法以集成其他类型的转换
class judgeFloat:
    def floatToInt(self,variable):
        variable="%d"%variable
        return variable
    def floatToStr(self,variable):
        variable=xlrd.xldate_as_tuple(variable,0)
        variable=list(variable)
        if variable[1]<10:
            variable[1]='0'+str(variable[1])
            variable=str(variable[0])+str(variable[1])+str(variable[2])
        return variable

#确定文档是否修改的函数:获取文档的修改时间与本地时间对比，不一致要求用户确认是否继续执行，一致则直接执行
def openFile(filename,address):
    filetime=time.strftime("%Y/%m/%d",time.localtime(os.stat(address+filename).st_mtime))
    local=time.strftime("%Y/%m/%d",time.localtime())
    if filetime!=local:
        single=input("文档今天还没有修改过，请确认是否继续？确认请输入1后点击回车，否则输入2点击回车。")
        if single=='1':
            pass
        elif single=='2':
            exit()
        else:
            print ('您的输入有误！脚本即将结束。')
            time.sleep(5)
            exit()
    else:
        pass

#写邮件的函数
def mailWrite(filename,address,i):
    header='<html><head><meta http-equiv="Content-Type" content="text/html; charset=utf-8" /></head>'
    th='<body text="#000000">committed详情：<table border="1" cellspacing="0" cellpadding="3" bordercolor="#000000" width="1800" align="left" ><tr bgcolor="#F79646" align="left" ><th>Purchaseing Document Number</th><th>Date on Which Record Was Create</th><th>Item Number of Purchasing Document</th><th>Short Text</th><th>Vendor Account Number</th><th>Vendor Name</th><th>Purchase Order Quantity</th><th>Purchase Order Unit of Measure</th><th>Currency Key</th><th>Net Price in Purchasing Document (in Doc</th><th>Price Unit</th><th>Quantity of Goods Received</th><th>Item Delivery Date</th><th>Text</th><th>Requester Email</th></tr>'
#打开文件
    filepath=address+filename
    book=xlrd.open_workbook(filepath)
    sheet=book.sheet_by_index(0)
#获取行列的数目，并以此为范围遍历获取单元数据
    nrows = sheet.nrows-1
    ncols = sheet.ncols
    body=''
    cellData=1
    td=''
    for j in range(ncols):
#读取单元格数据，赋给cellData变量供写入HTML表格中
        cellData=sheet.cell_value(i,j)
#调用浮点型转换方法解决读取的日期内容为浮点型数据的问题
        if isinstance(cellData,float):
            if (j==1 or j==12) and (i>0):
                cellDataNew=judgeFloat()
                cellData=cellDataNew.floatToStr(cellData)
            else:
                pass

        tip='<td>'+str(cellData)+'</td>'
#并入tr标签
        td=td+tip
        tr='<tr>'+td+'</tr>'
#为解决字符串拼接问题增设语句，tr从excel中读取出来是unicode编码，转换成UTF-8编码即可拼接
        tr=tr.encode('utf-8')
#并入body标签
    body=body+tr.decode()
    tail='</table></body></html>'
#将内容拼接成完整的HTML文档
    mail=header+th+body+tail
    return mail
#发送邮件------------------------------------------------------------------------
def mailSend(mail):
#循环每个发送者函数------------------------------------------------------------------------
   # for i in range(2,total+1):
       # r = sheet_ranges['O'+str(i)].value
        #设置接收人
        receiver = receivers[0]
#设置发件人
        sender = 'mic.test@magna.com'
#设置邮件主题
        subject = '测试邮件，请忽略！'
#设置发件服务器，即smtp服务器
        smtpserver = 'smtp.server.com'
#设置登陆名称
        username = 'server\mic.test'
#设置登陆密码
        password = '******'
#实例化写邮件到正文区，邮件正文区需要以HTML文档形式写入
        msg = MIMEText(mail,'html','utf-8')
#输入主题
        msg['Subject'] = subject
#调用邮件发送方法，需配合导入邮件相关模块
        smtp = smtplib.SMTP()
#设置连接发件服务器
        smtp.connect('smtp.server.com')
#输入用户名，密码，登陆服务器
        smtp.login(username, password)
#发送邮件
        smtp.sendmail(sender, receiver, msg.as_string())
#退出登陆并关闭与发件服务器的连接
        smtp.quit()
#入口函数，配置文件地址和文件名
def main():
    filename='file.xlsx'
    address='C:/File/'
    global r
    global receivers
    global i
    openFile(filename,address)
    for i in range(2,total+1):
        r = sheet_ranges['O'+str(i)].value
        mail=mailWrite(filename,address,i-1)
        receivers = [str(r)]
        #print(mail)
        #print(i)
        #print (receivers)
        mailSend(mail)
    
#调用执行main函数
if __name__=="__main__":
    main()